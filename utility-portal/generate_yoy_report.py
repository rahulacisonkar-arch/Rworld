import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import datetime
import re
import os
import json

# Get month argument
month_arg = 5 # default to May
is_json_mode = False

if len(sys.argv) > 1:
    try:
        month_arg = int(sys.argv[1])
    except:
        pass

if len(sys.argv) > 2 and sys.argv[2] == 'json':
    is_json_mode = True

months_map = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
    7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'
}
month_name = months_map.get(month_arg, 'May')

# Locate Excel file
file_path = r"C:\Users\Artee Admin\Downloads\TELE & INT EXP 2026 MAY MONTH .xlsx"
if not os.path.exists(file_path):
    print("Error: Excel source file not found!")
    sys.exit(1)

# Read file
df = pd.read_excel(file_path, sheet_name=0)

def parse_col_header(col):
    if isinstance(col, datetime.datetime):
        month = col.month
        if col.day in [23, 24, 25, 26]:
            year = 2000 + col.day
        else:
            year = col.year
        return month, year
    col_str = str(col).strip()
    match = re.match(r'^([A-Za-z]+)-(\d{2})(?:\.\d+)?$', col_str)
    if match:
        mon_name = match.group(1).upper()
        yr_short = int(match.group(2))
        year = 2000 + yr_short
        months_abbrev = {
            'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
            'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
        }
        if mon_name in months_abbrev:
            return months_abbrev[mon_name], year
    try:
        dt = pd.to_datetime(col_str)
        if pd.notna(dt) and dt.year > 100:
            month = dt.month
            if dt.day in [23, 24, 25, 26]:
                year = 2000 + dt.day
            else:
                year = dt.year
            return month, year
    except:
        pass
    return None, None

# Find columns for selected month
target_cols = []
for i, col in enumerate(df.columns):
    m, y = parse_col_header(col)
    if m == month_arg:
        target_cols.append((y, col, i))

# Years available
years = sorted(list(set([y for y, _, _ in target_cols])))
if not years:
    years = [2023, 2024, 2025, 2026] # fallback

# Process rows
rows_data = []
for idx, row in df.iterrows():
    code = row.get('Store code')
    if pd.isna(code):
        continue
    store_name = str(row.get('STORE')).strip()
    
    # Calculate sum for each year
    yoy_values = {}
    for y in years:
        yoy_values[y] = 0.0
        
    for y, col, i in target_cols:
        val = row.iloc[i]
        val_float = float(val) if pd.notna(val) and str(val).strip() != '' else 0.0
        yoy_values[y] = yoy_values.get(y, 0.0) + val_float
        
    record = {
        'code': int(float(code)),
        'store': store_name
    }
    for y in years:
        record[y] = yoy_values[y]
        
    rows_data.append(record)

# Create output folder inside portal
output_dir = r"C:\Users\Artee Admin\Desktop\browser-use-main\utility-portal\secure_uploads"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    
output_file = os.path.join(output_dir, f"YoY_Bill_Analysis_{month_name}.xlsx")

# Create a styled workbook using openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"{month_name} YoY Analysis"

# Styling definitions
font_family = "Segoe UI"
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy Blue
accent_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid") # Gold Accent
zebra_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Very light grey

font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
font_subtitle = Font(name=font_family, size=11, italic=True, color="E0E7FF")
font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
font_body = Font(name=font_family, size=10)
font_total = Font(name=font_family, size=10, bold=True)

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

double_bottom_border = Border(
    top=Side(style='thin', color='374151'),
    bottom=Side(style='double', color='374151')
)

# Header block
ws.merge_cells("A1:E1")
ws["A1"] = f"Year-over-Year Utility Analysis: {month_name.upper()}"
ws["A1"].font = font_title
ws["A1"].fill = header_fill
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:E2")
ws["A2"] = "Telephone & Internet Combined Expenses Report"
ws["A2"].font = font_subtitle
ws["A2"].fill = header_fill
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 20

# Column Headers
headers = ["Store Code", "Store Location"] + [f"{y} Combined" for y in years]
ws.row_dimensions[4].height = 25
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write Data Rows
current_row = 5
for idx, data in enumerate(rows_data):
    ws.row_dimensions[current_row].height = 20
    c1 = ws.cell(row=current_row, column=1, value=data['code'])
    c1.alignment = Alignment(horizontal="center")
    c1.font = font_body
    c1.border = thin_border
    
    c2 = ws.cell(row=current_row, column=2, value=data['store'])
    c2.font = font_body
    c2.border = thin_border
    
    for y_idx, y in enumerate(years, 3):
        val = data[y]
        c = ws.cell(row=current_row, column=y_idx, value=val)
        c.number_format = '$#,##0.00'
        c.alignment = Alignment(horizontal="right")
        c.font = font_body
        c.border = thin_border
        
    # Zebra striping
    if idx % 2 == 1:
        for c_idx in range(1, len(years) + 3):
            ws.cell(row=current_row, column=c_idx).fill = zebra_fill
            
    current_row += 1

# Total Row
ws.row_dimensions[current_row].height = 22
ws.cell(row=current_row, column=1, value="").border = double_bottom_border
t_label = ws.cell(row=current_row, column=2, value="TOTAL EXPENSE")
t_label.font = font_total
t_label.alignment = Alignment(horizontal="right")
t_label.border = double_bottom_border

for y_idx, y in enumerate(years, 3):
    col_letter = get_column_letter(y_idx)
    formula = f"=SUM({col_letter}5:{col_letter}{current_row-1})"
    c = ws.cell(row=current_row, column=y_idx, value=formula)
    c.number_format = '$#,##0.00'
    c.alignment = Alignment(horizontal="right")
    c.font = font_total
    c.border = double_bottom_border

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    if col[0].column > len(years) + 2:
        continue
    for cell in col:
        # Ignore merged cells in title
        if cell.row in [1, 2]:
            continue
        val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Insert YoY Chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = f"Telephone & Internet Combined Expenses - {month_name}"
chart.y_axis.title = "Combined Expense ($)"
chart.x_axis.title = "Retail Store Location"

# Years data reference
data_ref = Reference(ws, min_col=3, min_row=4, max_col=len(years)+2, max_row=current_row-1)
# Store name category reference
cats_ref = Reference(ws, min_col=2, min_row=5, max_row=current_row-1)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

chart.height = 16
chart.width = 25

ws.add_chart(chart, "G4")

if is_json_mode:
    # Print JSON output to stdout
    print(json.dumps({
        'years': years,
        'data': rows_data
    }))
    sys.exit(0)

# Save workbook
wb.save(output_file)
print(output_file)
