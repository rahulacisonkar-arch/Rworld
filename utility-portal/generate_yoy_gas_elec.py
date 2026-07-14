import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
import json
import re

# Get month argument
month_arg = 5 # default to May
is_json_mode = False

if len(sys.argv) > 1:
    try:
        month_arg = int(sys.argv[1])
    except:
        pass

if len(sys.argv) > 2 and sys.argv[2] == 'json':
    is_json_mode = True

months_map = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
    7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'
}
month_name = months_map.get(month_arg, 'May')

# Locate Excel file
file_path = r"C:\Users\Artee Admin\Downloads\store utility elec and gas analysis file.xlsx"
if not os.path.exists(file_path):
    print("Error: Excel source file not found!")
    sys.exit(1)

# Read file
df = pd.read_excel(file_path, sheet_name='Sheet1', header=None)

# 1. Map and parse columns using structured store-blocks of 10 columns each
active_stores = {
    '62': {'code': '62', 'name': 'Richmond'},
    '63': {'code': '63', 'name': 'Wilmington'},
    '64': {'code': '64', 'name': 'Virginia Beach'},
    '67': {'code': '67', 'name': 'Metairie'},
    '70': {'code': '70', 'name': 'Cincinnati'},
    '82': {'code': '82', 'name': 'Raleigh'},
    '83': {'code': '03', 'name': 'Burlington'}
}

years = [2024, 2025, 2026]
m_row_idx = 3 + month_arg
rows_data = []

# Map columns to store_code, store_name, year, utility
# Each store has 10 columns allocated starting at 1, 11, 21, ...
for k in range(13):
    c_start = 1 + k * 10
    if c_start >= df.shape[1]:
        break
        
    store_val = df.iloc[0, c_start]
    if pd.isna(store_val) or str(store_val).strip() == '':
        continue
        
    store_str = str(store_val).strip()
    match = re.search(r'-?\s*(\d+)$', store_str)
    if not match:
        continue
    excel_code = match.group(1)
    
    # Only map active stores
    if excel_code not in active_stores:
        continue
        
    store_info = active_stores[excel_code]
    store_name = store_info['name']
    store_code = store_info['code']
    
    yoy_values = {y: 0.0 for y in years}
    
    # Each store contains three triplets of columns corresponding to 2024, 2025, and 2026
    triplets = [
        (c_start, 2024),
        (c_start + 3, 2025),
        (c_start + 6, 2026)
    ]
    
    for start_col, default_year in triplets:
        if start_col >= df.shape[1]:
            continue
            
        # Determine the year for this triplet (look at Row 2 within the 3 columns)
        triplet_year = None
        for col_offset in range(3):
            col_idx = start_col + col_offset
            if col_idx < df.shape[1]:
                year_val = df.iloc[2, col_idx]
                if pd.notna(year_val) and str(year_val).strip() != '':
                    try:
                        triplet_year = int(float(year_val))
                        break
                    except:
                        pass
        if triplet_year is None:
            triplet_year = default_year
            
        # Parse utilities and sum values for Gas and Electricity
        for col_offset in range(3):
            col_idx = start_col + col_offset
            if col_idx < df.shape[1]:
                util_val = df.iloc[3, col_idx]
                if pd.notna(util_val) and str(util_val).strip() != '':
                    util_str = str(util_val).strip().lower()
                    if 'elect' in util_str or 'elec' in util_str:
                        utility = 'Electricity'
                    elif 'gas' in util_str:
                        utility = 'Gas'
                    else:
                        utility = None
                else:
                    utility = None
                    
                if utility in ['Gas', 'Electricity']:
                    val = df.iloc[m_row_idx, col_idx]
                    try:
                        val_float = float(val) if pd.notna(val) and str(val).strip() != '' else 0.0
                    except:
                        val_float = 0.0
                    yoy_values[triplet_year] = yoy_values.get(triplet_year, 0.0) + val_float
                    
    record = {
        'code': store_code,
        'store': store_name
    }
    for y in years:
        record[str(y)] = yoy_values[y]
    rows_data.append(record)

if is_json_mode:
    print(json.dumps({
        'years': [str(y) for y in years],
        'data': rows_data
    }))
    sys.exit(0)


# Create output folder inside portal
output_dir = r"C:\Users\Artee Admin\Desktop\browser-use-main\utility-portal\secure_uploads"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    
output_file = os.path.join(output_dir, f"YoY_Gas_Elec_Analysis_{month_name}.xlsx")

# Create a styled workbook using openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"{month_name} YoY Analysis"

font_family = "Segoe UI"
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy
zebra_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # light grey

font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
font_subtitle = Font(name=font_family, size=11, italic=True, color="E0E7FF")
font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
font_body = Font(name=font_family, size=10)
font_total = Font(name=font_family, size=10, bold=True)

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

double_bottom_border = Border(
    top=Side(style='thin', color='374151'),
    bottom=Side(style='double', color='374151')
)

# Header block
ws.merge_cells("A1:E1")
ws["A1"] = f"Year-over-Year Utility Analysis: {month_name.upper()}"
ws["A1"].font = font_title
ws["A1"].fill = header_fill
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:E2")
ws["A2"] = "Gas & Electricity Combined Expenses Report"
ws["A2"].font = font_subtitle
ws["A2"].fill = header_fill
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 20

# Column Headers
headers = ["Store Code", "Store Location"] + [f"{y} Combined" for y in years]
ws.row_dimensions[4].height = 25
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write Data Rows
current_row = 5
for idx, data in enumerate(rows_data):
    ws.row_dimensions[current_row].height = 20
    c1 = ws.cell(row=current_row, column=1, value=data['code'])
    c1.alignment = Alignment(horizontal="center")
    c1.font = font_body
    c1.border = thin_border
    
    c2 = ws.cell(row=current_row, column=2, value=data['store'])
    c2.font = font_body
    c2.border = thin_border
    
    for y_idx, y in enumerate(years, 3):
        val = data[str(y)]
        c = ws.cell(row=current_row, column=y_idx, value=val)
        c.number_format = '$#,##0.00'
        c.alignment = Alignment(horizontal="right")
        c.font = font_body
        c.border = thin_border
        
    if idx % 2 == 1:
        for c_idx in range(1, len(years) + 3):
            ws.cell(row=current_row, column=c_idx).fill = zebra_fill
            
    current_row += 1

# Total Row
ws.row_dimensions[current_row].height = 22
ws.cell(row=current_row, column=1, value="").border = double_bottom_border
t_label = ws.cell(row=current_row, column=2, value="TOTAL EXPENSE")
t_label.font = font_total
t_label.alignment = Alignment(horizontal="right")
t_label.border = double_bottom_border

for y_idx, y in enumerate(years, 3):
    col_letter = get_column_letter(y_idx)
    formula = f"=SUM({col_letter}5:{col_letter}{current_row-1})"
    c = ws.cell(row=current_row, column=y_idx, value=formula)
    c.number_format = '$#,##0.00'
    c.alignment = Alignment(horizontal="right")
    c.font = font_total
    c.border = double_bottom_border

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    if col[0].column > len(years) + 2:
        continue
    for cell in col:
        if cell.row in [1, 2]:
            continue
        val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Insert YoY Chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = f"Gas & Electricity Combined Expenses - {month_name}"
chart.y_axis.title = "Combined Expense ($)"
chart.x_axis.title = "Retail Store Location"

data_ref = Reference(ws, min_col=3, min_row=4, max_col=len(years)+2, max_row=current_row-1)
cats_ref = Reference(ws, min_col=2, min_row=5, max_row=current_row-1)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

chart.height = 16
chart.width = 25

ws.add_chart(chart, "G4")

wb.save(output_file)
print(output_file)
