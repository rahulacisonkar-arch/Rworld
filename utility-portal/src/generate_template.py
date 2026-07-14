import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utility Bills Import"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Store Code", "Utility Type", "Provider Name", 
        "Account Number", "Statement Date", "Due Date", 
        "Amount", "Status", "Transaction Reference", "Notes"
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Style header
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="164888", end_color="164888", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Sample data rows
    sample_data = [
        [53, "Telephone", "Verizon Business", "987654321-01", "2026-01-05", "2026-01-20", 145.50, "Paid", "TXN-98218A", "Monthly phone line fee (Atlanta-503)"],
        [83, "Internet", "Comcast Business", "84992019921", "2026-01-08", "2026-01-22", 189.99, "Pending", "", "High-speed broadband internet (Burlington-803)"],
        [70, "Electricity", "PSE&G", "0029-9118-28", "2026-01-10", "2026-01-25", 412.30, "Paid", "TXN-01992B", "Electrical utility (Cincinnati-700)"],
        [73, "Gas", "PSE&G NJ", "9918-2771-00", "2026-01-12", "2026-01-27", 325.40, "Pending", "", "Showroom gas heater supply (Good Goods-703)"],
        ["01", "Water", "Water Dept", "PW-99182", "2026-01-04", "2026-01-19", 68.20, "Paid", "TXN-88277C", "Water utility service (Ma'am)"],
        ["02", "Sewer", "Sewerage Authority", "SA-01992", "2026-01-05", "2026-01-20", 45.00, "Pending", "", "Sewer drainage maintenance (Warehouse)"]
    ]
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Write sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            
            # Alignments
            if col_idx in [5, 6]: # Dates
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 7: # Amount
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '$#,##0.00'
            elif col_idx == 8: # Status
                cell.alignment = Alignment(horizontal="center")
                # Highlight status
                if val == "Paid":
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, color="155724", bold=True)
                else:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, color="856404", bold=True)
            elif col_idx in [1, 4, 9]: # Codes/Numbers
                cell.alignment = Alignment(horizontal="center")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Format float/int for length check
                if isinstance(cell.value, float):
                    val_str = f"${cell.value:,.2f}"
                else:
                    val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Create public folder if not exists
    public_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "public")
    if not os.path.exists(public_dir):
        os.makedirs(public_dir)
        
    output_path = os.path.join(public_dir, "utility_upload_template.xlsx")
    wb.save(output_path)
    print(f"Template successfully generated at: {output_path}")

if __name__ == "__main__":
    create_template()
