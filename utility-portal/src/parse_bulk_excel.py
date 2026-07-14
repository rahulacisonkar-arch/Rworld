import openpyxl
import sys
import json
import os
import datetime

def parse_excel(file_path):
    if not os.path.exists(file_path):
        return {"success": False, "error": "File not found"}
        
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        return {"success": False, "error": f"Failed to load Excel workbook: {str(e)}"}
        
    rows = []
    errors = []
    
    # Headers check
    headers = []
    for col_idx in range(1, 11):
        headers.append(ws.cell(row=1, column=col_idx).value)
        
    expected_headers = [
        "Store Code", "Utility Type", "Provider Name", 
        "Account Number", "Statement Date", "Due Date", 
        "Amount", "Status", "Transaction Reference", "Notes"
    ]
    
    # Simple verification that it has the correct columns
    for idx, expected in enumerate(expected_headers):
        if not headers[idx] or expected.lower() not in str(headers[idx]).lower():
            return {
                "success": False, 
                "error": f"Header mismatch. Expected column {idx+1} to contain '{expected}', found '{headers[idx]}'"
            }
            
    # Read rows
    max_row = ws.max_row
    valid_utilities = ["Telephone", "Internet", "Gas", "Electricity", "Sewer", "Water"]
    
    for row_idx in range(2, max_row + 1):
        # Check if row is completely empty
        is_empty = True
        row_vals = []
        for col_idx in range(1, 11):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_vals.append(val)
            if val is not None and str(val).strip() != "":
                is_empty = False
                
        if is_empty:
            continue
            
        store_code, utility_type, provider_name, account_number, stmt_date, due_date, amount, status, txn_ref, notes = row_vals
        
        row_errs = []
        
        # 1. Validate Store Code
        if store_code is None or str(store_code).strip() == "":
            row_errs.append("Store Code is required.")
        else:
            if isinstance(store_code, float):
                if store_code.is_integer():
                    store_code = int(store_code)
            store_code_str = str(store_code).strip().upper()
            # If it's a float string like "53.0", strip the decimal
            if store_code_str.endswith(".0"):
                store_code_str = store_code_str[:-2]
            # Pad single digits (e.g. 1 -> "01", 2 -> "02")
            if store_code_str.isdigit() and len(store_code_str) == 1:
                store_code_str = "0" + store_code_str
            store_code = store_code_str
            
        # 2. Validate Utility Type
        if not utility_type or str(utility_type).strip() == "":
            row_errs.append("Utility Type is required.")
        else:
            utility_type = str(utility_type).strip().capitalize()
            if utility_type not in valid_utilities:
                row_errs.append(f"Invalid Utility Type: '{utility_type}'. Must be one of: {', '.join(valid_utilities)}")
                
        # 3. Provider Name
        if not provider_name or str(provider_name).strip() == "":
            row_errs.append("Provider Name is required.")
        else:
            provider_name = str(provider_name).strip()
            
        # 4. Account Number
        if not account_number or str(account_number).strip() == "":
            row_errs.append("Account Number is required.")
        else:
            account_number = str(account_number).strip()
            
        # 5 & 6. Validate Dates
        def parse_date(date_val, field_name):
            if isinstance(date_val, datetime.datetime):
                return date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, datetime.date):
                return date_val.strftime("%Y-%m-%d")
            
            if not date_val or str(date_val).strip() == "":
                row_errs.append(f"{field_name} is required.")
                return None
                
            date_str = str(date_val).strip()
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    dt = datetime.datetime.strptime(date_str, fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
            row_errs.append(f"Invalid {field_name} format: '{date_str}'. Use YYYY-MM-DD.")
            return None

        parsed_stmt_date = parse_date(stmt_date, "Statement Date")
        parsed_due_date = parse_date(due_date, "Due Date")
        
        # 7. Validate Amount
        parsed_amount = None
        if amount is None or str(amount).strip() == "":
            row_errs.append("Amount is required.")
        else:
            try:
                # Remove dollar sign, commas
                cleaned_amount = str(amount).replace('$', '').replace(',', '').strip()
                parsed_amount = float(cleaned_amount)
                if parsed_amount < 0:
                    row_errs.append("Amount cannot be negative.")
            except ValueError:
                row_errs.append(f"Invalid Amount: '{amount}'. Must be a numeric value.")
                
        # 8. Validate Status
        if not status or str(status).strip() == "":
            status = "Pending"
        else:
            status = str(status).strip().capitalize()
            if status not in ["Paid", "Pending", "Overdue"]:
                row_errs.append(f"Invalid Status: '{status}'. Must be 'Paid' or 'Pending'.")
                
        # Transaction Ref
        txn_ref = str(txn_ref).strip() if txn_ref else ""
        if status == "Paid" and txn_ref == "":
            row_errs.append("Transaction Reference is required if Status is 'Paid'.")
            
        # Notes
        notes = str(notes).strip() if notes else ""
        
        if row_errs:
            errors.append({
                "row": row_idx,
                "store_code": str(store_code) if store_code else "N/A",
                "utility_type": str(utility_type) if utility_type else "N/A",
                "reasons": row_errs
            })
        else:
            rows.append({
                "row_num": row_idx,
                "store_code": store_code,
                "utility_type": utility_type,
                "provider_name": provider_name,
                "account_number": account_number,
                "statement_date": parsed_stmt_date,
                "due_date": parsed_due_date,
                "amount": parsed_amount,
                "status": status,
                "transaction_ref": txn_ref,
                "notes": notes
            })
            
    return {
        "success": len(errors) == 0,
        "rows": rows,
        "errors": errors
    }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "No file path provided"}))
        sys.exit(1)
        
    target_file = sys.argv[1]
    result = parse_excel(target_file)
    print(json.dumps(result))
